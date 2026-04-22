def generate_attendance_excel(course_name, course_code, report_data, from_date, to_date) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Attendance Report'

        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f'Attendance Report - {course_name} ({course_code})'
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:H2')
        ws['A2'].value = f'Period: {from_date} to {to_date}'
        ws['A2'].alignment = Alignment(horizontal='center')

        headers = ['Roll No', 'Student Name', 'Total Classes',
                   'Present', 'Absent', 'Attendance %', 'Status', 'Emotion']
        
        for col_index, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_index)
            cell.value = h
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for i, student in enumerate(report_data):
            row_idx = 5 + i
            percentage = student.get('percentage', 0)
            if percentage >= 75:
                fill_color = '90EE90'
            elif percentage >= 60 and percentage < 75:
                fill_color = 'FFFF99'
            else:
                fill_color = 'FFB6C1'
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            if percentage >= 75:
                status_text = 'Regular'
            else:
                status_text = f'Defaulter (needs {student.get("shortfall",0)} more)'
            
            emotion_raw = student.get('last_emotion') or ''
            emotion_map = {
                'happy': 'Happy', 'sad': 'Sad', 'angry': 'Angry',
                'neutral': 'Neutral', 'surprised': 'Surprised',
                'fearful': 'Fearful', 'disgusted': 'Disgusted',
                'detecting': 'Neutral', '': 'N/A'
            }
            emotion_text = emotion_map.get(
                str(emotion_raw).lower().strip(), 
                str(emotion_raw) if emotion_raw else 'N/A'
            )
            
            values = [
                student.get('roll_number', ''), student.get('name', ''),
                student.get('total_classes', 0), student.get('present', 0),
                student.get('absent', 0), f"{percentage:.1f}%",
                status_text, emotion_text
            ]
            
            for col_index, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_index)
                cell.value = val
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise Exception(f'Excel generation failed: {str(e)}')